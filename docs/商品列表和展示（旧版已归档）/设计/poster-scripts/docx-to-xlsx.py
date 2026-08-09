"""Convert 价目表.docx into a clean 价目表.xlsx mirror.

Source layout (already inspected):
- 1 table, 61 rows, 3 columns.
- Column A in docx encodes the TV group (电视一…电视八) with stray newlines.
- 电视六 stuffs many drink rows into a single multi-line cell instead of one row each.
- 电视七 / 电视八 are a mix of group cells and trailing items with empty group column.

Output:
- A1:C1 merged title 「价目表」.
- Row 2 header: 电视编号 | 品名 | 价格.
- One row per item, 电视编号 vertically merged per group, 品名 left-aligned,
  价格 right-aligned, 电视编号 centered.
"""

from __future__ import annotations

import re
from pathlib import Path

from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "价目表.docx"
DST = ROOT / "价目表.xlsx"

TV_NORMALIZE = {
    "电视一": "电视一",
    "电视二": "电视二",
    "电视三": "电视三",
    "电视四": "电视四",
    "电视五": "电视五",
    "电视六": "电视六",
    "电视七": "电视七",
    "电视八": "电视八",
}


def normalize_tv(text: str) -> str | None:
    cleaned = re.sub(r"\s+", "", text or "")
    return TV_NORMALIZE.get(cleaned)


def split_inline_item(line: str) -> tuple[str, str] | None:
    """Split a free-form line like '可口可乐500ml  3.5' into (品名, 价格).

    Strategy: find the LAST run that looks like a price (number, optional 元/份/个/...),
    everything before it is the name.
    """
    s = line.strip()
    if not s:
        return None
    # Match trailing price tokens such as "3.5", "12元/份", "10元4个", "18/袋（半个）"
    m = re.search(
        r"([0-9]+(?:\.[0-9]+)?\s*(?:元)?(?:[/／](?:根|个|份|袋|盒|杯|斤|瓶|包|只|件|条))?(?:（[^）]*）)?)\s*$",
        s,
    )
    if not m:
        return s, ""
    price = m.group(1).strip()
    name = s[: m.start()].strip()
    if not name:
        return None
    return name, price


def extract_rows() -> list[tuple[str, str, str]]:
    """Return list of (电视编号, 品名, 价格) from the docx."""
    doc = Document(SRC)
    table = doc.tables[0]

    rows: list[tuple[str, str, str]] = []
    current_tv: str | None = None
    seen_pair: set[tuple[str, str, str]] = set()

    def push(tv: str, name: str, price: str) -> None:
        name = name.strip()
        price = price.strip()
        if not name:
            return
        key = (tv, name, price)
        if key in seen_pair:
            return
        seen_pair.add(key)
        rows.append(key)

    for ri, row in enumerate(table.rows):
        if ri == 0:
            continue  # 价目表 header
        if ri == 1:
            continue  # 电视一 / 品名 / 价格 header
        cells = [c.text for c in row.cells]
        col_a, col_b, col_c = cells[0], cells[1], cells[2]
        tv = normalize_tv(col_a)
        if tv:
            current_tv = tv
        if not current_tv:
            continue

        # Header rows have 品名/价格 literal in col_b/col_c -- skip.
        if col_b.strip() == "品名" and col_c.strip() == "价格":
            continue

        # Case 1: structured (品名 in col_b, 价格 in col_c) – may be multi-line.
        if col_b.strip() and (col_c.strip() or current_tv != "电视六"):
            name_lines = [ln.strip() for ln in col_b.splitlines() if ln.strip()]
            price_lines = [ln.strip() for ln in col_c.splitlines() if ln.strip()]
            # Common case: 1 name + 1 price.
            if len(name_lines) == 1 and len(price_lines) <= 1:
                push(current_tv, name_lines[0], price_lines[0] if price_lines else "")
                continue
            # 阜宁镇煎饼 has 1 name + 2 price lines → keep name + join prices with ";".
            if len(name_lines) == 1 and len(price_lines) > 1:
                push(current_tv, name_lines[0], "；".join(price_lines))
                continue
            # name_lines > 1: try to pair element-wise.
            if len(name_lines) == len(price_lines) and len(name_lines) > 1:
                for n, p in zip(name_lines, price_lines):
                    push(current_tv, n, p)
                continue
            # 电视六 weird case: col_b is a big blob with 品名+价格 mixed per line.
            for ln in name_lines:
                pair = split_inline_item(ln)
                if pair:
                    push(current_tv, pair[0], pair[1])
            for ln in price_lines:
                pair = split_inline_item(ln)
                if pair:
                    push(current_tv, pair[0], pair[1])
            continue

        # Case 2: only col_b filled (电视六 free-form rows like '拉罐可口可乐330ml 2.5').
        if col_b.strip() and not col_c.strip():
            for ln in col_b.splitlines():
                pair = split_inline_item(ln)
                if pair:
                    push(current_tv, pair[0], pair[1])

    return rows


def write_xlsx(rows: list[tuple[str, str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "价目表"

    title_font = Font(name="PingFang SC", size=18, bold=True, color="FFFFFF")
    header_font = Font(name="PingFang SC", size=12, bold=True, color="FFFFFF")
    body_font = Font(name="PingFang SC", size=12)
    tv_font = Font(name="PingFang SC", size=12, bold=True)

    title_fill = PatternFill("solid", fgColor="B91C1C")
    header_fill = PatternFill("solid", fgColor="DC2626")
    tv_fill = PatternFill("solid", fgColor="FEF2F2")

    thin = Side(style="thin", color="9CA3AF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    ws.merge_cells("A1:C1")
    ws["A1"] = "价目表"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 32

    headers = ["电视编号", "品名", "价格"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[2].height = 24

    start_row = 3
    for r, (tv, name, price) in enumerate(rows, start=start_row):
        a = ws.cell(row=r, column=1, value=tv)
        b = ws.cell(row=r, column=2, value=name)
        c = ws.cell(row=r, column=3, value=price)
        a.font = tv_font
        a.fill = tv_fill
        a.alignment = center
        b.font = body_font
        b.alignment = left
        c.font = body_font
        c.alignment = right
        for cell in (a, b, c):
            cell.border = border

    # Vertical merge per 电视编号 group.
    if rows:
        group_start = start_row
        for i in range(1, len(rows) + 1):
            cur_tv = rows[i - 1][0]
            next_tv = rows[i][0] if i < len(rows) else None
            if next_tv != cur_tv:
                end_row = start_row + i - 1
                if end_row > group_start:
                    ws.merge_cells(
                        start_row=group_start,
                        start_column=1,
                        end_row=end_row,
                        end_column=1,
                    )
                group_start = end_row + 1

    ws.column_dimensions[get_column_letter(1)].width = 12
    ws.column_dimensions[get_column_letter(2)].width = 32
    ws.column_dimensions[get_column_letter(3)].width = 18

    ws.freeze_panes = "A3"

    wb.save(DST)


def main() -> None:
    rows = extract_rows()
    write_xlsx(rows)
    by_tv: dict[str, int] = {}
    for tv, _, _ in rows:
        by_tv[tv] = by_tv.get(tv, 0) + 1
    print(f"Wrote {DST} with {len(rows)} rows.")
    for tv in sorted(by_tv):
        print(f"  {tv}: {by_tv[tv]} items")


if __name__ == "__main__":
    main()
